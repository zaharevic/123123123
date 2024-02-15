import datetime
import requests
import re
import openpyxl
import json
from io import StringIO
from model.classes.lesson import lesson
from model.classes.days import days
from model.parametrs.params import params
from model.parametrs.headers import headers


def parse_and_save_to_excel(data_string,
                            output_filename='C:/Users/Zahar/PycharmProjects/пары/парсинг/1/model/saved_data/tmp.xlsx'):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Data"
    records = data_string.strip().split('\n\n')
    headers1 = ["title", "teacher name", "start", "end"]
    for col, header in enumerate(headers1, start=1):
        sheet.cell(row=1, column=col, value=header)
    for record_idx, record in enumerate(records, start=2):
        record_file = StringIO(record)
        lines = record_file.readlines()
        for line in lines:
            split_line = line.strip().split(': ', 1)
            if len(split_line) == 2:
                key, value = split_line
                col = headers1.index(key) + 1
                sheet.cell(row=record_idx, column=col, value=value)
            else:
                print(f"Ignoring improperly formatted line: {line}")
    wb.save(output_filename)


class model:
    @staticmethod
    def get_calendar_info():
        url = "https://lk.samgtu.ru/api/common/distancelearning"

        response = requests.get(url, params=params, headers=headers)
        if response.ok:
            return response
        else:
            print(f"Error! Broken connection! Code: {response.status_code}")
            return 0

    @staticmethod
    def string_to_datatime(date, time='00:00'):
        string = date + ' ' + time
        return datetime.datetime.strptime(string, '%Y-%m-%d %H:%M')

    def get_day_rasp(self, json1, first_day, last_day):
        first_day_parsed = self.string_to_datatime(first_day)
        last_day_parsed = self.string_to_datatime(last_day, '23:59')
        flag = True
        days_tmp = days()
        i = 1
        while flag:
            tmp_data = json1.json()
            for dat in tmp_data:
                lesson_start = self.string_to_datatime(re.split('[,T]', dat['start'])[0],
                                                       re.split('[,T]', dat['start'])[1][:-3])
                if first_day_parsed <= lesson_start <= last_day_parsed:
                    lesson_name = dat['title']
                    lesson_end = self.string_to_datatime(re.split('[,T]', dat['end'])[0],
                                                         re.split('[,T]', dat['end'])[1][:-3])
                    teachers_name = dat['description'].split('<br>')[1]
                    lesson_tmp = lesson(lesson_name, lesson_start, lesson_end, teachers_name)
                    days_tmp.add(lesson_tmp)
                elif lesson_start > last_day_parsed:
                    flag = False
                    break
            i += 1
            flag = False
        return days_tmp

    @staticmethod
    def parse_and_save_to_json(data_string,
                               output_filename='C:/Users/Zahar/PycharmProjects/пары/парсинг/1/model/saved_data/data.json'):
        records = data_string.strip().split('\n\n')
        records_list = []
        for record in records:
            record_dict = {}
            lines = record.strip().split('\n')
            for line in lines:
                key, value = line.strip().split(': ', 1)
                record_dict[key] = value
            records_list.append(record_dict)
        with open(output_filename, 'w') as f:
            json.dump(records_list, f, ensure_ascii=False, indent=4)

    @staticmethod
    def print_rasp(data):
        print(data.__str__())

    def print_calendar(self):
        info = self.get_calendar_info()
        if info != 0:
            data = self.get_day_rasp(info, '2024-01-01', '2025-01-01')
            self.print_rasp(data)
