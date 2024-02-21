import datetime
from io import StringIO
import requests
import re
import json
import openpyxl
from model.clases.lesson import lesson
from model.clases.days import days
from model.parametrs.headers import headers
from model.parametrs.params import params
from model.clases.teachers_list import teachers_list
from model.clases.teacher import teacher
from model.clases.filtres import filtres


class core:
    def __init__(self):
        self.days_tmp = None
        self.teach_list = None
        self.filt = filtres()

    def start(self):
        a = self.get_calendar_info()
        self.start_parse(a)

    @staticmethod
    def get_calendar_info():
        url = "https://lk.samgtu.ru/api/common/distancelearning"
        response = requests.get(url, params=params, headers=headers)
        if response.ok:
            return response
        else:
            print(f"Error! Broken connection! Code: {response.status_code}")
            return 0

    @staticmethod
    def phpsessid_signing(phpsessid,
                          headers_file_path='C:/Users/Zahar/PycharmProjects/пары/парсинг/1/model/parametrs/headers.py'):
        url = 'https://lk.samgtu.ru/site/login'
        headers1 = {
            'Cookie': f'PHPSESSID={phpsessid}',
            'Referer': 'https://lk.samgtu.ru/distancelearning/distancelearning/index',
        }
        try:
            response = requests.get(url, headers=headers1)
            fake_resp = requests.get(url)
            if response.ok and len(response.text) != len(fake_resp.text):
                with open(headers_file_path, 'w') as f:
                    f.write("headers = {\n")
                    for key, value in headers1.items():
                        f.write(f"\t'{key}': '{value}',\n")
                    f.write("}\n")
                return True
            else:
                print(f'Error! Code:{response.status_code}')
                return False
        except Exception as e:
            print(f'Error occurred: {str(e)}')
            return True

    @staticmethod
    def string_to_datatime(date, time='00:00'):
        string = date + ' ' + time
        return datetime.datetime.strptime(string, '%Y-%m-%d %H:%M')

    def start_parse(self, json1):
        self.teach_list = teachers_list()
        self.unique_lessons = []
        flag = True
        self.days_tmp = days()
        i = 1

        while flag:
            tmp_data = json1.json()
            for dat in tmp_data:
                lesson_start = self.string_to_datatime(re.split('[,T]', dat['start'])[0],
                                                       re.split('[,T]', dat['start'])[1][:-3])
                lesson_name = dat['title']
                lesson_end = self.string_to_datatime(re.split('[,T]', dat['end'])[0],
                                                     re.split('[,T]', dat['end'])[1][:-3])
                try:
                    lesson_type = re.sub(r'<[^>]*>', '', dat['description'].split('<br>')[2])
                    teachers_name = dat['description'].split('<br>')[1].split(', ')
                except IndexError:
                    lesson_type = ''
                    teachers_name = []
                teachers_tmp = []
                for i in range(len(teachers_name)):
                    if self.teach_list.teacher_is_unique(teachers_name[i]):
                        id = self.teach_list.get_lenght()
                        teacher1 = teacher(id, teachers_name[i])
                        teachers_tmp.append(teacher1)
                        self.teach_list.append_teacher(teacher1)
                    else:
                        teachers_tmp.append(self.teach_list.get_teacher_by_name(teachers_name[i]))
                lesson_tmp = lesson(lesson_name, lesson_start, lesson_end, teachers_tmp, lesson_type)
                for t in teachers_tmp:
                    if lesson_name not in t.get_titles_less():
                        t.append_title(lesson_name)
                    if t is not None:
                        t.lesson_append(lesson_tmp)
                self.days_tmp.add(lesson_tmp)
                if lesson_name not in self.unique_lessons:
                    self.unique_lessons.append(lesson_name)
            i += 1
            flag = False
        return self.days_tmp

    def parse_and_save_to_excel(self, output_filename):
        data_string = self.days_tmp.str_for_save()
        if output_filename == '0':
            output_filename = 'C:/Users/Zahar/PycharmProjects/пары/парсинг/1/model/saved_data/tmp.xlsx'
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Data"
        records = data_string.strip().split('\n\n')
        headers1 = ["title", "teachers names", "start", "end", "lesson type"]
        for col, header in enumerate(headers1, start=1):
            sheet.cell(row=1, column=col, value=header)
        for record_idx, record in enumerate(records, start=2):
            record_file = StringIO(record)
            lines = record_file.readlines()
            for line in lines:
                split_line = line.strip().split(': ', 1)
                if len(split_line) == 2:
                    key, value = split_line
                    col = headers1.index(key) + 1
                    sheet.cell(row=record_idx, column=col, value=value)
                else:
                    print(f"Ignoring improperly formatted line: {line}")
        wb.save(output_filename)

    def parse_and_save_to_json(self, output_filename):
        data_string = self.days_tmp.__str__()
        if output_filename == '0':
            output_filename = 'C:/Users/Zahar/PycharmProjects/пары/парсинг/1/model/saved_data/data.json'
        records = data_string.strip().split('\n\n')
        records_list = []
        for record in records:
            record_dict = {}
            lines = record.strip().split('\n')
            for line in lines:
                key, value = line.strip().split(': ', 1)
                record_dict[key] = value
            records_list.append(record_dict)
        with open(output_filename, 'w') as f:
            json.dump(records_list, f, ensure_ascii=False, indent=4)

    def get_calendar(self, first_date, last_date):
        result = ''
        start_date = self.string_to_datatime(first_date, '00:00')
        last_date = self.string_to_datatime(last_date, '23:59')
        for less in self.days_tmp.get_lesson_list():
            if less.get_lesson_type() in self.filt.get_types_arr():
                if start_date < less.get_start() < last_date:
                    result += less.__str__()
                    result += '\n'
        return result

    def get_prepod_rasp(self, pr_id):
        prep = self.teach_list.get_teacher_by_id(pr_id)
        return prep.get_rasp()

    def try_parse_str_to_datatime(self, str1, str2):
        try:
            self.string_to_datatime(str1)
            self.string_to_datatime(str2)
            return True
        except Exception as e:
            print(f'Exception: {e}')
            return False

    def prep_is_real(self, id_pr):
        if self.teach_list.get_lenght() > id_pr:
            return True
        else:
            return False

    def switch_filter(self, code):
        if code == '1':
            self.filt.switch_lec()
        if code == '2':
            self.filt.switch_prac()
        if code == '3':
            self.filt.switch_lab()

    def get_filtres(self):
        return self.filt.__str__()

    def get_prep_list(self):
        return self.teach_list.__str__()

    def get_prep_titles(self, pr_id):
        t = self.teach_list.get_teacher_by_id(pr_id)
        return t.get_titles()

    def get_less_rasp(self, pr_id, l_id):
        t = self.teach_list.get_teacher_by_id(pr_id)
        arr = t.get_titles_less()
        return t.get_less_rasp(arr[l_id])

    def get_titles_lenght(self, pr_id):
        t = self.teach_list.get_teacher_by_id(pr_id)
        return len(t.get_titles_less())

    def get_prepod_info(self, le_id):
        res = ''
        teachres = self.get_prepods_by_less(le_id)
        for t in teachres:
            res += f'{t.get_info()} \n'
        return res

    def get_unique_less(self):
        res = 'Название предмета:\n'
        id1 = 0
        for less in self.unique_lessons:
            res += f'{id1}.{less}\n'
            id1 += 1
        return res

    def check_pred_id(self, id):
        if id < len(self.unique_lessons):
            return True
        else:
            return False

    def get_prepods_by_less(self, id):
        res = []
        pred = self.unique_lessons[id]
        for t in self.teach_list.get_teachers_list():
            if pred in t.get_titles_less():
                res.append(t)
        return res

    def print_less_rasp(self, less_id):
        less_name = self.unique_lessons[less_id]
        res = ''
        for less in self.days_tmp.get_lesson_list():
            if less.get_title() == less_name:
                res += less.__str__()
                res += '\n'
        return res
